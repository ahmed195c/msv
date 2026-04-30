import calendar
import datetime
import io
import os

from django.contrib.auth import login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Group, User
from django.http import HttpResponse
from django.utils import timezone
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Count, Q
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse

from ..models import (
    Company, CompanyChangeLog, EngineerCertificateRequest, EngineerLeave,
    Enginer, EnginerStatusLog, InspectorReview, PesticideTransportPermit,
    PirmetChangeLog, PirmetClearance, PirmetDocument, PublicHealthExamRequest,
    PublicHealthExamRequestDocument, RequirementInsuranceRequest,
    WasteDisposalRequest, WasteDisposalRequestDocument,
)
from ..forms import StaffRegistrationForm
from .common import (
    ALLOWED_DOC_EXTENSIONS, PEST_ACTIVITY_ORDER, PEST_ACTIVITY_KEYS,
    PUBLIC_HEALTH_ACTIVITY_KEYS, GROUP_NAME_ALIASES, ROLE_CAPABILITIES,
    INSPECTION_REPORT_PHOTO_PREFIX, VEHICLE_INSPECTION_REPORT_PHOTO_PREFIX,
    _parse_int, _parse_int_list, _parse_date, _calculate_permit_expiry,
    _add_months, _expired_trade_license_notice, _activities_for_enginer,
    _restricted_activities_for_enginer, _has_any_group, _role_is_admin,
    _role_is_inspector, _role_is_data_entry, _role_is_head, _user_roles,
    _has_capability, _can_admin, _can_inspector, _can_data_entry, _can_head,
    _company_has_active_extension, _can_create_exam_request,
    _inspector_users_qs, _display_user_name, _inspector_review_name,
    _inspection_report_decision_from_note, _inspection_report_photo_count_from_note,
    _inspection_report_photo_docs_by_prefix, _inspection_report_photo_docs,
    _vehicle_inspection_report_photo_docs, _request_documents,
    _latest_expired_activity_permit_before, _delay_months_after_first_month,
    _initial_violation_reference_expiry, _violation_reference_expiry_date,
    _log_pirmet_change, _log_company_change, _split_activities,
    _activity_keys_for_company, _permit_label_ar, _permit_detail_url_name,
    _certificate_type_for_exam, _certificate_expiry, _enginer_has_passed_for_certificate,
    _is_effective_active_permit, _engineer_no_certificate_notice,
    _group_clearances_by_status, _validate_engineer_for_type,
)
@login_required
def clearance_list(request):
    search_query = (request.GET.get('q') or '').strip()
    clearances_qs = (
        PirmetClearance.objects.filter(permit_type__in=['pest_control', 'pesticide_transport', 'waste_disposal', 'engineer_addition'])
        .select_related('company', 'company__enginer')
        .order_by('-dateOfCreation')
    )
    if search_query:
        clearances_qs = clearances_qs.filter(
            Q(company__name__icontains=search_query)
            | Q(company__number__icontains=search_query)
        )
    reviews = InspectorReview.objects.filter(pirmet__in=clearances_qs).select_related('inspector', 'inspector_user')
    review_map = {review.pirmet_id: review for review in reviews}
    inspection_receive_changes = (
        PirmetChangeLog.objects.filter(
            pirmet__in=clearances_qs,
            change_type='details_update',
            notes__startswith='inspection_received_by:',
        )
        .order_by('pirmet_id', '-created_at')
    )
    inspection_report_changes = (
        PirmetChangeLog.objects.filter(
            pirmet__in=clearances_qs,
            change_type='details_update',
            notes__startswith='inspection_report:',
        )
        .select_related('changed_by')
        .order_by('pirmet_id', '-created_at')
    )
    inspection_receive_map = {}
    for change in inspection_receive_changes:
        if change.pirmet_id not in inspection_receive_map:
            inspection_receive_map[change.pirmet_id] = change
    inspection_report_map = {}
    for change in inspection_report_changes:
        if change.pirmet_id not in inspection_report_map:
            inspection_report_map[change.pirmet_id] = change
    clearances = list(clearances_qs)
    waste_requests_qs = (
        WasteDisposalRequest.objects.filter(permit__in=clearances_qs)
        .select_related('inspected_by')
        .order_by('permit_id', '-created_at', '-id')
    )
    latest_waste_request_map = {}
    latest_active_waste_request_map = {}
    for waste_request in waste_requests_qs:
        if waste_request.permit_id not in latest_waste_request_map:
            latest_waste_request_map[waste_request.permit_id] = waste_request
        if (
            waste_request.status in {'payment_pending', 'inspection_pending'}
            and waste_request.permit_id not in latest_active_waste_request_map
        ):
            latest_active_waste_request_map[waste_request.permit_id] = waste_request

    for clearance in clearances:
        clearance.permit_label_ar = _permit_label_ar(clearance.permit_type)
        clearance.detail_url_name = _permit_detail_url_name(clearance.permit_type)
        clearance.detail_url = reverse(clearance.detail_url_name, kwargs={'id': clearance.id})
        clearance.active_waste_request = None
        if clearance.permit_type == 'waste_disposal':
            latest_waste_request = (
                latest_active_waste_request_map.get(clearance.id)
                or latest_waste_request_map.get(clearance.id)
            )
            if latest_waste_request:
                clearance.permit_label_ar = 'طلب التخلص من النفايات'
                clearance.detail_url = reverse(
                    'waste_disposal_request_detail',
                    kwargs={'permit_id': clearance.id, 'request_id': latest_waste_request.id},
                )
            clearance.active_waste_request = latest_active_waste_request_map.get(clearance.id)
        company = clearance.company
        engineer = company.enginer if company else None
        engineer_phone = (engineer.phone or '').strip() if engineer else ''
        owner_phone = (company.owner_phone or '').strip() if company else ''
        landline_phone = (company.landline or '').strip() if company else ''
        clearance.contact_number = engineer_phone or owner_phone or landline_phone or None
        clearance.company_location = (company.address or '').strip() if company and company.address else '-'
        review = review_map.get(clearance.id)
        clearance.inspector_name = _inspector_review_name(review)
        clearance.inspection_assigned_user_id = review.inspector_user_id if review and review.inspector_user_id else None
        if (
            not clearance.inspection_assigned_user_id
            and clearance.permit_type == 'waste_disposal'
        ):
            pending_waste_request = latest_active_waste_request_map.get(clearance.id)
            if (
                pending_waste_request
                and pending_waste_request.status == 'inspection_pending'
                and pending_waste_request.inspected_by_id
            ):
                clearance.inspection_assigned_user_id = pending_waste_request.inspected_by_id
        request_docs = _request_documents(clearance)
        clearance.request_documents_count = len(request_docs)
        clearance.has_request_documents = bool(
            clearance.request_documents_bundle
            or clearance.request_documents_count
        )
        receive_change = inspection_receive_map.get(clearance.id)
        report_change_item = inspection_report_map.get(clearance.id)
        clearance.inspection_receiver_name = None
        if receive_change and ':' in receive_change.notes:
            clearance.inspection_receiver_name = receive_change.notes.split(':', 1)[1].strip()
        # Compute inspection duration for admin indicator
        clearance.inspection_duration_hours = None
        clearance.inspection_duration_display = None
        if receive_change and report_change_item:
            delta = report_change_item.created_at - receive_change.created_at
            hours = max(int(delta.total_seconds() // 3600), 0)
            clearance.inspection_duration_hours = hours
            clearance.inspection_duration_display = f'{hours // 24} يوم' if hours >= 24 else f'{hours} ساعة'
        elif receive_change and clearance.status == 'inspection_pending':
            delta = timezone.now() - receive_change.created_at
            hours = max(int(delta.total_seconds() // 3600), 0)
            clearance.inspection_duration_hours = hours
            clearance.inspection_duration_display = f'{hours // 24} يوم' if hours >= 24 else f'{hours} ساعة'
        if (
            not clearance.inspection_receiver_name
            and clearance.permit_type == 'waste_disposal'
        ):
            pending_waste_request = latest_active_waste_request_map.get(clearance.id)
            if pending_waste_request and pending_waste_request.inspected_by:
                clearance.inspection_receiver_name = _display_user_name(
                    pending_waste_request.inspected_by
                )
        clearance.status_key = clearance.status
        if clearance.active_waste_request:
            wr_status = clearance.active_waste_request.status
            if wr_status == 'payment_pending':
                clearance.status_key = 'waste_disposal_payment_pending'
            elif wr_status == 'inspection_pending':
                clearance.status_key = 'waste_disposal_inspection_pending'
        elif (
            clearance.status == 'inspection_pending'
            and clearance.inspection_receiver_name
        ):
            clearance.status_key = 'inspection_received'
        report_change = inspection_report_map.get(clearance.id)
        clearance.inspection_report_decision = (
            _inspection_report_decision_from_note(report_change.notes)
            if report_change
            else None
        )
        if not clearance.inspection_report_decision and clearance.permit_type == 'waste_disposal':
            if clearance.status == 'disposal_approved':
                clearance.inspection_report_decision = 'approved'
            elif clearance.status == 'disposal_rejected':
                clearance.inspection_report_decision = 'rejected'
        if (
            clearance.status == 'inspection_completed'
            and clearance.inspection_report_decision == 'approved'
            and clearance.permit_type == 'pesticide_transport'
        ):
            clearance.status_key = 'pesticide_payment_link_pending'

    finished_statuses = {
        'issued',
        'closed_requirements_pending',
        'cancelled_admin',
        'disposal_approved',
        'disposal_rejected',
    }
    active_clearances = []
    finished_clearances = []
    for _item in clearances:
        has_active_waste_request = (
            _item.permit_type == 'waste_disposal'
            and _item.id in latest_active_waste_request_map
        )
        if has_active_waste_request:
            active_clearances.append(_item)
        elif _item.status in finished_statuses:
            finished_clearances.append(_item)
        elif (
            _item.status == 'inspection_completed'
            and _item.inspection_report_decision != 'approved'
        ):
            finished_clearances.append(_item)
        else:
            active_clearances.append(_item)
    inspector_scope_only = _can_inspector(request.user) and not _can_admin(request.user)

    if _can_inspector(request.user):
        current_user_id = request.user.id

        def _active_inspector_sort_key(item):
            item_status = getattr(item, 'status_key', item.status)
            is_ready_to_receive = item_status == 'inspection_pending'
            is_received = item_status == 'inspection_received'
            assigned_to_current = bool(
                is_received
                and item.inspection_assigned_user_id
                and item.inspection_assigned_user_id == current_user_id
            )
            if is_ready_to_receive:
                priority = 0
            elif assigned_to_current:
                priority = 1
            elif is_received:
                priority = 2
            else:
                priority = 3
            created_ordinal = item.dateOfCreation.toordinal() if item.dateOfCreation else 0
            return (priority, -created_ordinal, -item.id)

        active_clearances.sort(key=_active_inspector_sort_key)
    elif _can_admin(request.user):
        def _active_admin_sort_key(item):
            item_status = getattr(item, 'status_key', item.status)
            needs_head_approval = (
                item_status == 'inspection_completed'
                and getattr(item, 'inspection_report_decision', None) == 'approved'
            )
            priority = 0 if needs_head_approval else 1
            created_ordinal = item.dateOfCreation.toordinal() if item.dateOfCreation else 0
            return (priority, -created_ordinal, -item.id)
        active_clearances.sort(key=_active_admin_sort_key)
    else:
        active_clearances.sort(
            key=lambda item: (
                -(item.dateOfCreation.toordinal() if item.dateOfCreation else 0),
                -item.id,
            )
        )

    finished_ids = [item.id for item in finished_clearances]
    finished_completion_map = {}
    if finished_ids:
        completion_logs = (
            PirmetChangeLog.objects.filter(
                pirmet_id__in=finished_ids,
                change_type='status_change',
            )
            .order_by('pirmet_id', '-created_at')
        )
        for log in completion_logs:
            if log.pirmet_id not in finished_completion_map:
                finished_completion_map[log.pirmet_id] = log.created_at
    for item in finished_clearances:
        item._completion_date = finished_completion_map.get(item.id)
    finished_clearances.sort(
        key=lambda item: (
            -(item._completion_date.timestamp() if item._completion_date else 0),
            -item.id,
        )
    )

    status_filter_label_map = {
        'waste_disposal_payment_pending':   'بانتظار دفع طلب الإتلاف',
        'waste_disposal_inspection_pending': 'طلب إتلاف جاهز للاستلام',
        'inspection_received': 'تم استلام الطلب للتفتيش',
        'inspection_pending': 'جاهز للاستلام',
        'order_received': 'بانتظار اصدار رابط دفع التفتيش',
        'inspection_payment_pending': 'بانتظار دفع التفتيش',
        'review_pending': 'بانتظار مراجعة المفتش',
        'approved': 'معتمد من المفتش',
        'needs_completion': 'غير معتمد',
        'rejected': 'مرفوض',
        'payment_pending': 'بانتظار دفع التصريح',
        'issued': 'تم إصدار التصريح',
        'inspection_completed': 'اكتمل التفتيش',
        'pesticide_payment_link_pending': 'تصريح المركبة - بانتظار إرسال رابط الدفع',
        'violation_payment_link_pending': 'بانتظار إرسال رابط دفع المخالفة',
        'violation_payment_pending': 'بانتظار دفع المخالفة',
        'closed_requirements_pending': 'مغلق - اشتراطات واجبة الاستيفاء',
        'cancelled_admin': 'مغلق',
        'disposal_approved': 'إتلاف معتمد',
        'disposal_rejected': 'إتلاف مرفوض',
    }
    status_section_label_map = {
        'waste_disposal_payment_pending':   'طلبات إتلاف بانتظار الدفع',
        'waste_disposal_inspection_pending': 'طلبات إتلاف جاهزة للاستلام',
        'inspection_received': 'طلبات تم استلامها للتفتيش',
        'inspection_pending': 'طلبات جاهزة للاستلام',
        'order_received': 'بانتظار اصدار رابط دفع التفتيش',
        'inspection_payment_pending': 'طلبات بانتظار دفع التفتيش',
        'review_pending': 'طلبات بانتظار مراجعة المفتش',
        'approved': 'طلبات معتمدة من المفتش',
        'needs_completion': 'طلبات غير معتمدة',
        'rejected': 'طلبات مرفوضة',
        'payment_pending': 'طلبات بانتظار دفع التصريح',
        'violation_payment_link_pending': 'طلبات بانتظار إرسال رابط دفع المخالفة',
        'violation_payment_pending': 'طلبات بانتظار دفع المخالفة',
        'issued': 'طلبات صادرة',
        'inspection_completed': 'طلبات اكتمل تفتيشها',
        'pesticide_payment_link_pending': 'تصاريح المركبات - بانتظار إرسال رابط الدفع',
        'closed_requirements_pending': 'طلبات مغلقة - اشتراطات واجبة الاستيفاء',
        'cancelled_admin': 'طلبات مغلقة',
        'disposal_approved': 'طلبات إتلاف معتمدة',
        'disposal_rejected': 'طلبات إتلاف مرفوضة',
    }
    active_status_order = [
        'waste_disposal_payment_pending',
        'waste_disposal_inspection_pending',
        'inspection_completed',
        'approved',
        'head_approved',
        'inspection_pending',
        'inspection_received',
        'order_received',
        'inspection_payment_pending',
        'review_pending',
        'payment_pending',
        'pesticide_payment_link_pending',
        'violation_payment_link_pending',
        'violation_payment_pending',
        'needs_completion',
        'rejected',
    ]
    finished_status_order = [
        'issued',
        'inspection_completed',
        'closed_requirements_pending',
        'cancelled_admin',
        'disposal_approved',
        'disposal_rejected',
    ]
    if inspector_scope_only:
        active_status_order = ['inspection_pending', 'inspection_received']
        finished_status_order = []
    all_status_order = active_status_order + finished_status_order

    status_filter = (request.GET.get('status') or 'all').strip()
    if status_filter != 'all' and status_filter not in status_filter_label_map:
        status_filter = 'all'

    if status_filter != 'all':
        active_clearances = [
            item for item in active_clearances
            if getattr(item, 'status_key', item.status) == status_filter
        ]
        finished_clearances = [
            item for item in finished_clearances
            if getattr(item, 'status_key', item.status) == status_filter
        ]

    status_filter_options = [
        {'value': 'all', 'label': 'كل الحالات'},
        *[
            {'value': status_key, 'label': status_filter_label_map.get(status_key, status_key)}
            for status_key in all_status_order
        ],
    ]
    status_filter_label = (
        status_filter_label_map.get(status_filter, 'كل الحالات')
        if status_filter != 'all'
        else 'كل الحالات'
    )

    active_clearance_groups = _group_clearances_by_status(active_clearances, active_status_order, status_section_label_map)
    finished_clearance_groups = _group_clearances_by_status(finished_clearances, finished_status_order, status_section_label_map)

    _permit_type_tabs = [
        ('pest_control',        'تصاريح مزاولة النشاط'),
        ('pesticide_transport', 'تصاريح المركبة'),
        ('waste_disposal',      'تصاريح التخلص من النفايات'),
        ('engineer_addition',   'طلبات إضافة مهندس'),
    ]
    _valid_tab_keys = {k for k, _ in _permit_type_tabs}
    active_tab = (request.GET.get('tab') or '').strip()
    if active_tab not in _valid_tab_keys:
        active_tab = 'pest_control'

    permit_type_tab_data = []
    for pt_key, pt_label in _permit_type_tabs:
        tab_active = [c for c in active_clearances if c.permit_type == pt_key]
        tab_finished = [c for c in finished_clearances if c.permit_type == pt_key]
        permit_type_tab_data.append({
            'key': pt_key,
            'label': pt_label,
            'active_clearances': tab_active,
            'finished_clearances': tab_finished,
            'active_groups': _group_clearances_by_status(tab_active, active_status_order, status_section_label_map),
            'finished_groups': _group_clearances_by_status(tab_finished, finished_status_order, status_section_label_map),
            'active_count': len(tab_active),
        })

    return render(
        request,
        'hcsd/clearance_list.html',
        {
            'clearances': active_clearances,
            'active_clearances': active_clearances,
            'finished_clearances': finished_clearances,
            'active_clearance_groups': active_clearance_groups,
            'finished_clearance_groups': finished_clearance_groups,
            'permit_type_tab_data': permit_type_tab_data,
            'active_tab': active_tab,
            'query': search_query,
            'status_filter': status_filter,
            'status_filter_label': status_filter_label,
            'status_filter_options': status_filter_options,
            'show_finished_section': not inspector_scope_only,
            'can_create_pirmet': _can_data_entry(request.user),
            'user_is_admin': _can_admin(request.user),
            'form_errors': [],
        },
    )


@login_required
def permit_types(request):
    today = timezone.localdate()
    week_later = today + datetime.timedelta(days=7)

    permit_label_map = {
        'pest_control': 'تصريح مزاولة النشاط',
        'pesticide_transport': 'تصريح المركبة',
        'waste_disposal': 'تصريح التخلص من النفايات',
    }

    def _latest_per_company_type(qs):
        """Keep only the newest permit per (company, permit_type) pair."""
        seen = {}
        for p in qs:
            key = (p.company_id, p.permit_type)
            if key not in seen:
                seen[key] = p
        return list(seen.values())

    def _enrich(permits):
        result = []
        for p in permits:
            engineer = p.company.enginer if p.company else None
            phone = (engineer.phone or '').strip() if engineer else ''
            if not phone and p.company:
                phone = (p.company.owner_phone or p.company.landline or '').strip()
            p.contact_phone = phone or '—'
            p.permit_label = permit_label_map.get(p.permit_type, p.permit_type)
            result.append(p)
        return result

    # Fetch all issued permits ordered newest first so _latest_per_company_type keeps the newest
    all_issued = list(
        PirmetClearance.objects
        .select_related('company', 'company__enginer')
        .filter(status='issued')
        .order_by('-issue_date', '-id')
    )
    latest_issued = _latest_per_company_type(all_issued)

    expiring_permits = []
    for p in latest_issued:
        if p.dateOfExpiry and today <= p.dateOfExpiry <= week_later:
            p.days_left = (p.dateOfExpiry - today).days
            expiring_permits.append(p)
    expiring_permits.sort(key=lambda p: p.dateOfExpiry)

    finished_permits = sorted(
        [p for p in latest_issued if p.dateOfExpiry and p.dateOfExpiry < today],
        key=lambda p: p.dateOfExpiry,
        reverse=True,
    )[:50]

    return render(
        request,
        'hcsd/permit_types.html',
        {
            'can_create_pirmet': _can_data_entry(request.user),
            'expiring_permits': _enrich(expiring_permits),
            'finished_permits': _enrich(finished_permits),
        },
    )


@login_required
def permits_report_excel(request):
    """Export one row per company with permit status columns (check/cross + permit no)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    today = timezone.localdate()

    ACTIVITY_LABEL = {
        'public_health_pest_control': 'مكافحة آفات الصحة العامة',
        'termite_control': 'مكافحة النمل الأبيض',
        'grain_pests': 'مكافحة آفات الحبوب',
    }

    # Fetch all companies with related engineer
    companies = list(
        Company.objects.select_related('enginer')
    )
    company_ids = [c.id for c in companies]

    # Fetch all permits ordered: issued first, then newest creation date
    all_permits = list(
        PirmetClearance.objects
        .filter(company_id__in=company_ids)
        .order_by('company_id', 'permit_type', '-dateOfCreation', '-id')
    )

    # Build map: company_id -> {permit_type -> permit}
    # Prefer the latest *issued* permit over any in-progress permit
    permit_map = {}
    for p in all_permits:
        cid = p.company_id
        pt = p.permit_type
        if cid not in permit_map:
            permit_map[cid] = {}
        existing = permit_map[cid].get(pt)
        if existing is None:
            permit_map[cid][pt] = p
        elif existing.status != 'issued' and p.status == 'issued':
            # Replace in-progress permit with an issued one
            permit_map[cid][pt] = p

    # Sort companies: newest pest_control permit first, companies with no permit last
    def _company_sort_key(company):
        pc = permit_map.get(company.id, {}).get('pest_control')
        if pc:
            return (0, -pc.dateOfCreation.toordinal())
        return (1, 0)

    companies.sort(key=_company_sort_key)

    # Active leave map for all engineers
    all_eng_ids = {c.enginer_id for c in companies if c.enginer_id}
    active_leave_map = {
        leave.engineer_id: leave
        for leave in EngineerLeave.objects.filter(
            actual_return_date__isnull=True,
            engineer_id__in=all_eng_ids,
        ).select_related('substitute')
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'تقرير الشركات والتصاريح'
    ws.sheet_view.rightToLeft = True

    # ── Styles ──
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='304357')
    sub_header_fill = PatternFill('solid', fgColor='455c75')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True, readingOrder=2)
    right_align = Alignment(horizontal='right', vertical='center', wrap_text=True, readingOrder=2)
    thin = Side(border_style='thin', color='C0C8D0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    green_fill = PatternFill('solid', fgColor='D1FAE5')
    red_fill   = PatternFill('solid', fgColor='FEE2E2')
    green_font = Font(bold=True, color='065F46', size=13)
    red_font   = Font(bold=True, color='991B1B', size=13)

    # ── Headers ──
    # Row 1: group headers
    # Row 2: column headers
    PERMIT_TYPES = [
        ('pest_control',       'تصريح النشاط'),
        ('pesticide_transport','تصريح المركبة'),
        ('waste_disposal',     'تصريح النفايات'),
    ]

    # columns: م | اسم الشركة | رقم الرخصة | العنوان | المنطقة | الهاتف | المهندس | حالة المهندس
    #          | [for each permit type: الحالة | رقم التصريح | تاريخ الانتهاء]
    info_cols = ['م', 'اسم الشركة', 'رقم الرخصة', 'العنوان', 'هاتف الشركة', 'المهندس', 'حالة المهندس', 'الأنشطة المرخصة']
    n_info = len(info_cols)
    permit_sub_cols = ['الحالة', 'رقم التصريح', 'تاريخ الانتهاء']
    n_permit_cols = len(permit_sub_cols)

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 24

    # Row 1: info group header (merge) + permit group headers (merge each)
    # Merge info cols in row 1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_info)
    cell = ws.cell(row=1, column=1, value='بيانات الشركة')
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border

    for ti, (pt_key, pt_label) in enumerate(PERMIT_TYPES):
        start_col = n_info + ti * n_permit_cols + 1
        end_col   = start_col + n_permit_cols - 1
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        cell = ws.cell(row=1, column=start_col, value=pt_label)
        cell.font = header_font
        cell.fill = sub_header_fill
        cell.alignment = center
        cell.border = border

    # Row 2: individual column headers
    for col_idx, label in enumerate(info_cols, 1):
        cell = ws.cell(row=2, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for ti in range(len(PERMIT_TYPES)):
        for si, sub_label in enumerate(permit_sub_cols):
            col_idx = n_info + ti * n_permit_cols + si + 1
            cell = ws.cell(row=2, column=col_idx, value=sub_label)
            cell.font = header_font
            cell.fill = sub_header_fill
            cell.alignment = center
            cell.border = border

    # Column widths
    col_widths = [4, 32, 16, 32, 14, 24, 14, 30]  # info cols
    for _ in PERMIT_TYPES:
        col_widths += [8, 14, 14]                  # status, permit_no, expiry
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Data rows ──
    row_num = 3
    for idx, company in enumerate(companies, 1):
        engineer = company.enginer
        eng_name = engineer.name if engineer else ''
        if engineer:
            active_leave = active_leave_map.get(engineer.id)
            eng_status = 'في إجازة' if active_leave else 'في العمل'
        else:
            eng_status = ''

        phone = ''
        if engineer:
            phone = (engineer.phone or '').strip()
        if not phone:
            phone = (company.owner_phone or company.landline or '').strip()

        # Get allowed activities from the latest active pest_control permit
        active_pc_permit = permit_map.get(company.id, {}).get('pest_control')
        activities_display = ''
        if active_pc_permit and active_pc_permit.allowed_activities:
            keys = [k.strip() for k in active_pc_permit.allowed_activities.split(',') if k.strip()]
            activities_display = ' / '.join(
                ACTIVITY_LABEL.get(k, k) for k in keys
            )

        info_data = [
            idx,
            company.name or '',
            company.number or '',
            company.address or '',
            phone,
            eng_name,
            eng_status,
            activities_display,
        ]

        ws.row_dimensions[row_num].height = 20
        for col_idx, value in enumerate(info_data, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.alignment = center if col_idx == 1 else right_align
            cell.border = border

        company_permits = permit_map.get(company.id, {})
        for ti, (pt_key, _) in enumerate(PERMIT_TYPES):
            p = company_permits.get(pt_key)
            base_col = n_info + ti * n_permit_cols + 1

            if p is None:
                # No permit at all
                status_val = '✗'
                permit_no_val = ''
                expiry_val = ''
                s_fill = red_fill
                s_font = red_font
            else:
                is_active = (
                    p.status == 'issued'
                    and (not p.dateOfExpiry or p.dateOfExpiry >= today)
                )
                status_val = '✓' if is_active else '✗'
                permit_no_val = p.permit_no or ''
                expiry_val = p.dateOfExpiry.strftime('%d/%m/%Y') if p.dateOfExpiry else ''
                s_fill = green_fill if is_active else red_fill
                s_font = green_font if is_active else red_font

            # Status cell (✓ / ✗)
            c = ws.cell(row=row_num, column=base_col, value=status_val)
            c.font = s_font
            c.fill = s_fill
            c.alignment = center
            c.border = border

            # Permit number
            c2 = ws.cell(row=row_num, column=base_col + 1, value=permit_no_val)
            c2.alignment = center
            c2.border = border

            # Expiry date
            c3 = ws.cell(row=row_num, column=base_col + 2, value=expiry_val)
            c3.alignment = center
            c3.border = border

        row_num += 1

    ws.freeze_panes = 'B3'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"companies_report_{today.strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def inspector_report_excel(request):
    """Export inspector performance report: rows = inspectors, cols = permit type × decision."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    today = timezone.localdate()

    PERMIT_TYPES = [
        ('pest_control',        'تصريح النشاط'),
        ('pesticide_transport', 'تصريح المركبة'),
        ('waste_disposal',      'تصريح النفايات'),
    ]
    PT_KEYS = [k for k, _ in PERMIT_TYPES]

    # Only count permits where inspection is actually completed
    COMPLETED_STATUSES = {
        'inspection_completed', 'head_approved', 'approved',
        'violation_payment_link_pending', 'violation_payment_pending',
        'payment_pending', 'issued', 'cancelled_admin',
        'closed_requirements_pending', 'needs_completion',
    }

    # Fetch all reviews with related data (completed permits only)
    reviews = list(
        InspectorReview.objects
        .select_related('inspector_user', 'pirmet')
        .filter(
            inspector_user__isnull=False,
            pirmet__status__in=COMPLETED_STATUSES,
        )
    )

    # Fetch change-log decisions for all permits (new-flow stores decision here)
    from collections import defaultdict
    pirmet_ids = [r.pirmet_id for r in reviews if r.pirmet_id]
    log_decisions = {}
    for log in (
        PirmetChangeLog.objects
        .filter(
            pirmet_id__in=pirmet_ids,
            change_type='details_update',
            notes__startswith='inspection_report:',
        )
        .order_by('pirmet_id', '-created_at')
        .values('pirmet_id', 'notes')
    ):
        pid = log['pirmet_id']
        if pid not in log_decisions:
            dec = log['notes'].split(':', 1)[1].strip().lower()
            if dec in {'approved', 'rejected', 'requirements_required'}:
                log_decisions[pid] = dec

    # Build: {user_id: {permit_type: {approved: N, rejected: N}}}
    stats = defaultdict(lambda: {pt: {'approved': 0, 'rejected': 0} for pt in PT_KEYS})
    user_map = {}

    for r in reviews:
        uid = r.inspector_user_id
        pt = r.pirmet.permit_type if r.pirmet_id else None
        if pt not in PT_KEYS:
            continue
        user_map[uid] = r.inspector_user

        # New flow: decision stored in change log (accurate)
        # Old flow: no change log entry → fall back to isApproved
        log_dec = log_decisions.get(r.pirmet_id)
        if log_dec == 'approved':
            decision = 'approved'
        elif log_dec in ('rejected', 'requirements_required'):
            decision = 'rejected'
        elif r.isApproved:
            decision = 'approved'
        else:
            decision = 'rejected'

        stats[uid][pt][decision] += 1

    # Sort inspectors by total reviews desc
    inspector_ids = sorted(
        user_map.keys(),
        key=lambda uid: sum(
            stats[uid][pt]['approved'] + stats[uid][pt]['rejected']
            for pt in PT_KEYS
        ),
        reverse=True,
    )

    # ── Styles ──
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'تقرير المفتشين'
    ws.sheet_view.rightToLeft = True

    header_font   = Font(bold=True, color='FFFFFF', size=11)
    header_fill   = PatternFill('solid', fgColor='1e293b')
    sub_fill_ap   = PatternFill('solid', fgColor='166534')
    sub_fill_rj   = PatternFill('solid', fgColor='991b1b')
    sub_fill_tot  = PatternFill('solid', fgColor='374151')
    center        = Alignment(horizontal='center', vertical='center', wrap_text=True, readingOrder=2)
    right_align   = Alignment(horizontal='right',  vertical='center', wrap_text=True, readingOrder=2)
    thin          = Side(border_style='thin', color='C0C8D0')
    border        = Border(left=thin, right=thin, top=thin, bottom=thin)

    green_fill = PatternFill('solid', fgColor='D1FAE5')
    red_fill   = PatternFill('solid', fgColor='FEE2E2')
    green_font = Font(bold=True, color='065F46')
    red_font   = Font(bold=True, color='991B1B')
    gray_fill  = PatternFill('solid', fgColor='F3F4F6')
    gray_font  = Font(bold=True, color='374151')

    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 22

    # ── Row 1: group headers ──
    # Info group
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    c = ws.cell(row=1, column=1, value='بيانات المفتش')
    c.font = header_font; c.fill = header_fill; c.alignment = center; c.border = border

    # Permit type groups (3 cols each: اعتماد | رفض | المجموع)
    for ti, (_, pt_label) in enumerate(PERMIT_TYPES):
        sc = 3 + ti * 3
        ws.merge_cells(start_row=1, start_column=sc, end_row=1, end_column=sc + 2)
        c = ws.cell(row=1, column=sc, value=pt_label)
        c.font = header_font; c.fill = header_fill; c.alignment = center; c.border = border

    # Total group
    total_col = 3 + len(PERMIT_TYPES) * 3
    ws.merge_cells(start_row=1, start_column=total_col, end_row=1, end_column=total_col + 2)
    c = ws.cell(row=1, column=total_col, value='الإجمالي')
    c.font = header_font; c.fill = header_fill; c.alignment = center; c.border = border

    # ── Row 2: sub-headers ──
    for ci, label in enumerate(['م', 'اسم المفتش'], 1):
        c = ws.cell(row=2, column=ci, value=label)
        c.font = header_font; c.fill = header_fill; c.alignment = center; c.border = border

    sub_labels = ['اعتماد', 'رفض', 'المجموع']
    for ti in range(len(PERMIT_TYPES)):
        for si, sl in enumerate(sub_labels):
            col = 3 + ti * 3 + si
            c = ws.cell(row=2, column=col, value=sl)
            c.font = header_font
            c.fill = sub_fill_ap if si == 0 else (sub_fill_rj if si == 1 else sub_fill_tot)
            c.alignment = center; c.border = border

    for si, sl in enumerate(sub_labels):
        col = total_col + si
        c = ws.cell(row=2, column=col, value=sl)
        c.font = header_font
        c.fill = sub_fill_ap if si == 0 else (sub_fill_rj if si == 1 else sub_fill_tot)
        c.alignment = center; c.border = border

    # ── Column widths ──
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 26
    for ti in range(len(PERMIT_TYPES)):
        for si in range(3):
            ws.column_dimensions[get_column_letter(3 + ti * 3 + si)].width = 10
    for si in range(3):
        ws.column_dimensions[get_column_letter(total_col + si)].width = 10

    # ── Data rows ──
    row_num = 3
    total_approved_all = 0
    total_rejected_all = 0

    for idx, uid in enumerate(inspector_ids, 1):
        user = user_map[uid]
        name = user.get_full_name() or user.username
        ws.row_dimensions[row_num].height = 19

        ws.cell(row=row_num, column=1, value=idx).alignment = center
        ws.cell(row=row_num, column=1).border = border
        ws.cell(row=row_num, column=2, value=name).alignment = right_align
        ws.cell(row=row_num, column=2).border = border

        row_approved = 0
        row_rejected = 0

        for ti, (pt_key, _) in enumerate(PERMIT_TYPES):
            ap = stats[uid][pt_key]['approved']
            rj = stats[uid][pt_key]['rejected']
            tot = ap + rj
            row_approved += ap
            row_rejected += rj

            base = 3 + ti * 3
            ca = ws.cell(row=row_num, column=base, value=ap or '')
            ca.alignment = center; ca.border = border
            if ap: ca.fill = green_fill; ca.font = green_font

            cr = ws.cell(row=row_num, column=base + 1, value=rj or '')
            cr.alignment = center; cr.border = border
            if rj: cr.fill = red_fill; cr.font = red_font

            ct = ws.cell(row=row_num, column=base + 2, value=tot or '')
            ct.alignment = center; ct.border = border
            if tot: ct.fill = gray_fill; ct.font = gray_font

        total_approved_all += row_approved
        total_rejected_all += row_rejected

        ca2 = ws.cell(row=row_num, column=total_col, value=row_approved or '')
        ca2.alignment = center; ca2.border = border
        if row_approved: ca2.fill = green_fill; ca2.font = green_font

        cr2 = ws.cell(row=row_num, column=total_col + 1, value=row_rejected or '')
        cr2.alignment = center; cr2.border = border
        if row_rejected: cr2.fill = red_fill; cr2.font = red_font

        ct2 = ws.cell(row=row_num, column=total_col + 2, value=(row_approved + row_rejected) or '')
        ct2.alignment = center; ct2.border = border
        if row_approved + row_rejected: ct2.fill = gray_fill; ct2.font = gray_font

        row_num += 1

    # ── Totals row ──
    ws.row_dimensions[row_num].height = 20
    total_font = Font(bold=True, size=11)
    total_fill = PatternFill('solid', fgColor='E2E8F0')

    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
    ct = ws.cell(row=row_num, column=1, value='الإجمالي')
    ct.font = total_font; ct.fill = total_fill; ct.alignment = center; ct.border = border

    grand_approved = 0
    grand_rejected = 0
    for ti, (pt_key, _) in enumerate(PERMIT_TYPES):
        ap_total = sum(stats[uid][pt_key]['approved'] for uid in inspector_ids)
        rj_total = sum(stats[uid][pt_key]['rejected'] for uid in inspector_ids)
        grand_approved += ap_total
        grand_rejected += rj_total
        base = 3 + ti * 3

        c = ws.cell(row=row_num, column=base, value=ap_total or '')
        c.font = total_font; c.fill = total_fill; c.alignment = center; c.border = border
        c = ws.cell(row=row_num, column=base + 1, value=rj_total or '')
        c.font = total_font; c.fill = total_fill; c.alignment = center; c.border = border
        c = ws.cell(row=row_num, column=base + 2, value=(ap_total + rj_total) or '')
        c.font = total_font; c.fill = total_fill; c.alignment = center; c.border = border

    c = ws.cell(row=row_num, column=total_col, value=grand_approved or '')
    c.font = total_font; c.fill = total_fill; c.alignment = center; c.border = border
    c = ws.cell(row=row_num, column=total_col + 1, value=grand_rejected or '')
    c.font = total_font; c.fill = total_fill; c.alignment = center; c.border = border
    c = ws.cell(row=row_num, column=total_col + 2, value=(grand_approved + grand_rejected) or '')
    c.font = total_font; c.fill = total_fill; c.alignment = center; c.border = border

    ws.freeze_panes = 'C3'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"inspector_report_{today.strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
