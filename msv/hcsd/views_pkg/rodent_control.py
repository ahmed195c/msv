"""
Rodent Control — building trap (RBS) monitoring.

URL prefix : /rodent-control/
Templates  : hcsd/rodent_control_*.html

One trap per building. A monthly visit record is auto-generated for every
active building on the 1st of each month (see
management/commands/generate_rodent_control_visits.py); staff fill it in
when the team actually visits.
"""

import datetime
import io

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST

from ..models import RodentControlBuilding, RodentControlVisit
from .common import _can_admin, _can_data_entry


def _can_manage(user):
    return _can_admin(user) or _can_data_entry(user)


def _current_period_start(today=None):
    today = today or timezone.localdate()
    return today.replace(day=1)


def _get_or_create_current_visit(building, today=None):
    period_start = _current_period_start(today)
    visit, _ = RodentControlVisit.objects.get_or_create(
        building=building, period_start=period_start,
    )
    return visit


@login_required
def rodent_control_list(request):
    query = (request.GET.get('q') or '').strip()

    buildings_qs = RodentControlBuilding.objects.filter(is_active=True)
    if query:
        buildings_qs = buildings_qs.filter(name__icontains=query)
    buildings = list(buildings_qs.order_by('name'))

    period_start = _current_period_start()
    current_visits = {
        v.building_id: v
        for v in RodentControlVisit.objects.filter(
            building__in=buildings, period_start=period_start,
        )
    }

    rows = []
    for b in buildings:
        visit = current_visits.get(b.id)
        rows.append({'building': b, 'visit': visit})

    return render(request, 'hcsd/rodent_control_list.html', {
        'rows': rows,
        'query': query,
        'period_start': period_start,
        'can_manage': _can_manage(request.user),
        'total_buildings': len(buildings),
    })


@login_required
def rodent_control_building_create(request):
    if not _can_manage(request.user):
        return redirect('rodent_control_list')

    errors = []
    if request.method == 'POST':
        name     = (request.POST.get('name') or '').strip()
        number   = (request.POST.get('number') or '').strip()
        area     = (request.POST.get('area') or '').strip()
        location = (request.POST.get('location') or '').strip()
        notes    = (request.POST.get('notes') or '').strip()

        if not name:
            errors.append('يرجى إدخال اسم البناية.')

        if not errors:
            building = RodentControlBuilding.objects.create(
                name=name, number=number, area=area, location=location,
                notes=notes, created_by=request.user,
            )
            _get_or_create_current_visit(building)
            return redirect('rodent_control_building_detail', pk=building.pk)

    return render(request, 'hcsd/rodent_control_building_create.html', {
        'errors': errors,
        'post': request.POST,
    })


@login_required
def rodent_control_building_detail(request, pk):
    building = get_object_or_404(RodentControlBuilding, pk=pk)
    can_manage = _can_manage(request.user)

    current_visit = _get_or_create_current_visit(building)
    history = list(
        building.visits.exclude(pk=current_visit.pk).order_by('-period_start')
    )

    if request.method == 'POST':
        if not can_manage:
            return redirect('rodent_control_building_detail', pk=pk)

        action = request.POST.get('action', '')

        if action == 'record_visit':
            def _int(name):
                raw = (request.POST.get(name) or '').strip()
                try:
                    return int(raw) if raw else None
                except ValueError:
                    return None

            def _float(name):
                raw = (request.POST.get(name) or '').strip()
                try:
                    return float(raw) if raw else None
                except ValueError:
                    return None

            def _time(name):
                raw = (request.POST.get(name) or '').strip()
                try:
                    return datetime.time.fromisoformat(raw) if raw else None
                except ValueError:
                    return None

            visit_date_raw = (request.POST.get('visit_date') or '').strip()
            try:
                visit_date = datetime.date.fromisoformat(visit_date_raw)
            except ValueError:
                visit_date = timezone.localdate()

            current_visit.visit_date = visit_date
            current_visit.visited_by = request.user

            current_visit.team_leader_name = (request.POST.get('team_leader_name') or '').strip()
            current_visit.team_leader_id = (request.POST.get('team_leader_id') or '').strip()
            current_visit.time_in = _time('time_in')
            current_visit.time_out = _time('time_out')

            current_visit.rbs_inspected_count = _int('rbs_inspected_count')
            current_visit.rbs_lock_ok = 'rbs_lock_ok' in request.POST
            current_visit.rbs_infested_count = _int('rbs_infested_count')
            current_visit.rbs_damaged_count = _int('rbs_damaged_count')
            current_visit.rbs_new_installed_count = _int('rbs_new_installed_count')
            current_visit.stick_change_ok = 'stick_change_ok' in request.POST
            current_visit.rbs_replenished_count = _int('rbs_replenished_count')

            current_visit.manholes_inspected_count = _int('manholes_inspected_count')
            current_visit.manholes_treated_count = _int('manholes_treated_count')
            current_visit.manholes_treated_qty = _float('manholes_treated_qty')
            current_visit.manholes_infested_count = _int('manholes_infested_count')

            current_visit.burrows_outside_count = _int('burrows_outside_count')
            current_visit.burrows_infested_count = _int('burrows_infested_count')

            current_visit.trees_inspected_count = _int('trees_inspected_count')
            current_visit.trees_treated_count = _int('trees_treated_count')
            current_visit.trees_infested_count = _int('trees_infested_count')

            current_visit.rodenticide_type = (request.POST.get('rodenticide_type') or '').strip()
            current_visit.rodenticide_quantity = _float('rodenticide_quantity')
            current_visit.notes = (request.POST.get('notes') or '').strip()

            # Auto-derive the summary flags (used by the list-page badge and
            # by the historical import) from the actual counts just entered.
            current_visit.inspected = bool(current_visit.rbs_inspected_count)
            current_visit.infested = bool(current_visit.rbs_infested_count) or bool(current_visit.manholes_infested_count) or bool(current_visit.burrows_infested_count) or bool(current_visit.trees_infested_count)
            current_visit.damaged = bool(current_visit.rbs_damaged_count) or not current_visit.rbs_lock_ok
            current_visit.newly_installed = bool(current_visit.rbs_new_installed_count)
            current_visit.replenished = bool(current_visit.rbs_replenished_count)

            current_visit.save()
            return redirect('rodent_control_building_detail', pk=pk)

        elif action == 'update_building':
            building.name     = (request.POST.get('name') or building.name).strip()
            building.number   = (request.POST.get('number') or '').strip()
            building.area     = (request.POST.get('area') or '').strip()
            building.location = (request.POST.get('location') or '').strip()
            building.notes    = (request.POST.get('notes') or '').strip()
            building.save(update_fields=['name', 'number', 'area', 'location', 'notes'])
            return redirect('rodent_control_building_detail', pk=pk)

        elif action == 'toggle_active':
            building.is_active = not building.is_active
            building.save(update_fields=['is_active'])
            return redirect('rodent_control_building_detail', pk=pk)

    return render(request, 'hcsd/rodent_control_building_detail.html', {
        'building': building,
        'current_visit': current_visit,
        'history': history,
        'can_manage': can_manage,
    })


@login_required
def rodent_control_monthly_excel(request):
    """Excel export of visit records for a date range, in the same layout
    used by the old hand-built monthly spreadsheets — so future reports can
    be produced straight from this system instead of being rebuilt by hand."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    today = timezone.localdate()
    raw_from = (request.GET.get('date_from') or '').strip()
    raw_to = (request.GET.get('date_to') or '').strip()
    try:
        date_from = datetime.date.fromisoformat(raw_from)
    except ValueError:
        date_from = today.replace(day=1)
    try:
        date_to = datetime.date.fromisoformat(raw_to)
    except ValueError:
        date_to = today
    if date_from > date_to:
        date_from, date_to = date_to, date_from

    visits = list(
        RodentControlVisit.objects.filter(
            period_start__gte=date_from.replace(day=1),
            period_start__lte=date_to,
        )
        .select_related('building', 'visited_by')
        .order_by('building__name', 'period_start')
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'تقرير المصايد'
    ws.sheet_view.rightToLeft = True

    thin = Side(style='thin', color='999999')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill('solid', fgColor='0e7490')
    ok_fill = PatternFill('solid', fgColor='e8f5ee')
    bad_fill = PatternFill('solid', fgColor='fdeaea')

    headers = [
        '#', 'اسم البناية', 'المنطقة', 'الشهر', 'تاريخ الزيارة',
        'قائد الفريق', 'الرقم الوظيفي', 'دخول', 'خروج',
        'مصايد مفتشة', 'قفل سليم', 'مصايد مصابة', 'مصايد تالفة',
        'تركيب جديد', 'تغيير لاصقة', 'مصايد معبأة',
        'مناهيل مفتشة', 'مناهيل معالجة (عدد)', 'مناهيل معالجة (كمية)', 'مناهيل مصابة',
        'جحور خارجية', 'جحور مصابة',
        'نخيل مفتش', 'نخيل معالج', 'نخيل مصاب',
        'نوع المادة', 'الكمية', 'ملاحظات',
    ]
    widths = [5, 26, 14, 9, 12, 16, 10, 8, 8, 10, 8, 10, 10, 9, 10, 10, 10, 10, 10, 10, 10, 10, 9, 9, 9, 22, 9, 26]
    for col, (hdr, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=1, column=col, value=hdr)
        c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 30

    def _n(v):
        return v if v is not None else '—'

    for row_idx, v in enumerate(visits, start=2):
        visited_by_name = v.team_leader_name or (
            v.visited_by.get_full_name() or v.visited_by.username if v.visited_by else '—'
        )
        values = [
            row_idx - 1,
            v.building.name,
            v.building.area or '—',
            v.period_start.strftime('%m/%Y'),
            v.visit_date.strftime('%d/%m/%Y') if v.visit_date else '—',
            visited_by_name,
            v.team_leader_id or '—',
            v.time_in.strftime('%H:%M') if v.time_in else '—',
            v.time_out.strftime('%H:%M') if v.time_out else '—',
            _n(v.rbs_inspected_count),
            'نعم' if v.rbs_lock_ok else 'لا',
            _n(v.rbs_infested_count),
            _n(v.rbs_damaged_count),
            _n(v.rbs_new_installed_count),
            'نعم' if v.stick_change_ok else 'لا',
            _n(v.rbs_replenished_count),
            _n(v.manholes_inspected_count),
            _n(v.manholes_treated_count),
            _n(v.manholes_treated_qty),
            _n(v.manholes_infested_count),
            _n(v.burrows_outside_count),
            _n(v.burrows_infested_count),
            _n(v.trees_inspected_count),
            _n(v.trees_treated_count),
            _n(v.trees_infested_count),
            v.rodenticide_type or '—',
            v.rodenticide_quantity if v.rodenticide_quantity is not None else '—',
            v.notes or '—',
        ]
        for col, val in enumerate(values, start=1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.font = Font(name='Arial', size=10)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = border
            header_key = headers[col - 1]
            if header_key == 'مصايد مصابة' and v.rbs_infested_count:
                c.fill = bad_fill
            elif header_key == 'مصايد تالفة' and v.rbs_damaged_count:
                c.fill = bad_fill
            elif header_key == 'مصايد مفتشة' and v.rbs_inspected_count:
                c.fill = ok_fill

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'rodent_control_{date_from:%Y-%m}_{date_to:%Y-%m}.xlsx'
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
