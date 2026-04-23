"""
Management command to import field work orders from the Excel summary file.

Usage:
    python manage.py seed_field_work
    python manage.py seed_field_work --file path/to/file.xlsx
    python manage.py seed_field_work --clear     # delete existing excel records first
    python manage.py seed_field_work --sheet "APR - 26"  # import one sheet only
"""

import datetime
import os

from django.core.management.base import BaseCommand, CommandError

from hcsd.models import FieldWorkOrder

EXCEL_PATH_DEFAULT = os.path.join(
    os.path.dirname(__file__),
    '..', '..', 'static', 'hcsd', 'excl',
    'Summary of Unscheduled 2026 (1).xlsx',
)

MONTH_SHEETS = ['JAN - 26', 'FEB - 26', 'MAR - 26', 'APR - 26']

# Col indices (0-based)
COL_ROW_NUM    = 0
COL_ORDER_NUM  = 1
COL_REQ_DATE   = 2
COL_CLOSE_DATE = 3
COL_CUSTOMER   = 4
COL_STATUS     = 5
COL_STATUS2    = 6
COL_MOBILE     = 7
COL_STREET     = 8
COL_HOUSE      = 9
COL_AREA       = 10
COL_PESTS      = 11
COL_SUPERVISOR = 12
COL_WORKER     = 13
COL_ANT        = 14
COL_COCKROACH  = 15
COL_MOSQUITO   = 16
# col 17 = باعوض (duplicate mosquito, skipped)
COL_FLY        = 18
COL_RAT        = 19
COL_SNAKE      = 20
COL_SCORPION   = 21
COL_WASPS      = 22
COL_BEES       = 23
COL_OTHER      = 24
COL_BOOM       = 25
COL_KOTHRENI   = 26
COL_DIESEL     = 27
COL_PETROL     = 28
COL_CYPHORCE   = 29
COL_RAT_POISON = 30
COL_ECO_LARV   = 31
COL_SNAKE_DET  = 32
COL_HYMENOP    = 33
COL_PERMOTHOR  = 34
COL_RAT_GLUE   = 35
COL_RAPETR     = 36
COL_GRAIBAIT   = 37
COL_DIFRON     = 38
COL_FLY_ATTR   = 39

COMPLETED_STATUSES = {'THE SERVICE HAS BEEN COMPLETED'}
CANCELLED_STATUSES = {
    'THE CUSTOMER DECLINED',
    'PRIVATE COMPANY',
    'BELONGING TO OTHER MUNICIPALITIES',
    'GOVERNMENT DEPARTMENT, APPROVAL MUST BE SENT',
}


def _bool(val):
    return bool(val and str(val).strip())


def _str(val, maxlen=None):
    s = str(val).strip() if val is not None else ''
    if maxlen:
        s = s[:maxlen]
    return s


def _date(val):
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    return None


def _map_status(excel_status):
    s = (excel_status or '').strip().upper()
    if s in COMPLETED_STATUSES:
        return 'completed'
    if s in CANCELLED_STATUSES:
        return 'cancelled'
    return 'pending'


class Command(BaseCommand):
    help = 'Import field work orders from the Excel summary file'

    def add_arguments(self, parser):
        parser.add_argument('--file', default=None, help='Path to Excel file')
        parser.add_argument('--sheet', default=None, help='Import specific sheet only')
        parser.add_argument(
            '--clear', action='store_true',
            help='Delete all existing Excel-imported records before importing',
        )
        parser.add_argument(
            '--skip-existing', action='store_true', default=True,
            help='Skip rows whose order_number already exists (default: True)',
        )
        parser.add_argument(
            '--update-existing', action='store_true',
            help='Update existing records instead of skipping',
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            raise CommandError('openpyxl is required: pip install openpyxl')

        file_path = options['file'] or os.path.normpath(EXCEL_PATH_DEFAULT)
        if not os.path.exists(file_path):
            raise CommandError(f'File not found: {file_path}')

        if options['clear']:
            deleted, _ = FieldWorkOrder.objects.filter(source='excel').delete()
            self.stdout.write(self.style.WARNING(f'Deleted {deleted} existing Excel records.'))

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        sheets = [options['sheet']] if options['sheet'] else MONTH_SHEETS
        for sheet_name in sheets:
            if sheet_name not in wb.sheetnames:
                self.stdout.write(self.style.WARNING(f'Sheet "{sheet_name}" not found, skipping.'))
                continue
            self._import_sheet(wb[sheet_name], sheet_name, options)

        self.stdout.write(self.style.SUCCESS('Import complete.'))

    def _import_sheet(self, ws, sheet_name, options):
        update_existing = options.get('update_existing', False)
        created = updated = skipped = 0

        rows = list(ws.iter_rows(values_only=True))
        for row in rows[1:]:  # skip header
            # Only require a valid order number — skip truly empty rows
            if not row[COL_ORDER_NUM]:
                continue

            order_num = _str(row[COL_ORDER_NUM], 30)
            excel_status = _str(row[COL_STATUS], 100)
            excel_status2 = _str(row[COL_STATUS2], 100)

            defaults = dict(
                source='excel',
                month_sheet=sheet_name,
                request_date=_date(row[COL_REQ_DATE]),
                close_date=_date(row[COL_CLOSE_DATE]),
                customer_name=_str(row[COL_CUSTOMER], 200),
                excel_status=excel_status,
                excel_status_note=excel_status2,
                status=_map_status(excel_status),
                mobile=_str(row[COL_MOBILE], 30),
                street_number=_str(row[COL_STREET], 50),
                house_number=_str(row[COL_HOUSE], 50),
                area=_str(row[COL_AREA], 200),
                pest_types=_str(row[COL_PESTS], 300),
                supervisor_name=_str(row[COL_SUPERVISOR], 200),
                worker_name=_str(row[COL_WORKER], 200),
                # Pest treatments
                treated_ant=_bool(row[COL_ANT]),
                treated_cockroach=_bool(row[COL_COCKROACH]),
                treated_mosquito=_bool(row[COL_MOSQUITO]),
                treated_fly=_bool(row[COL_FLY]),
                treated_rat=_bool(row[COL_RAT]),
                treated_snake=_bool(row[COL_SNAKE]),
                treated_scorpion=_bool(row[COL_SCORPION]),
                treated_wasps=_bool(row[COL_WASPS]),
                treated_bees=_bool(row[COL_BEES]),
                treated_other=_bool(row[COL_OTHER]),
                # Materials
                used_boom=_bool(row[COL_BOOM]),
                used_kothreni=_bool(row[COL_KOTHRENI]),
                used_diesel=_bool(row[COL_DIESEL]),
                used_petrol=_bool(row[COL_PETROL]),
                used_cyphorce=_bool(row[COL_CYPHORCE]),
                used_rat_poison=_bool(row[COL_RAT_POISON]),
                used_eco_larvacide=_bool(row[COL_ECO_LARV]),
                used_snake_deter=_bool(row[COL_SNAKE_DET]),
                used_hymenopthor=_bool(row[COL_HYMENOP]),
                used_permothor=_bool(row[COL_PERMOTHOR]),
                used_rat_glue=_bool(row[COL_RAT_GLUE]),
                used_rapetr_gel=_bool(row[COL_RAPETR]),
                used_graibait=_bool(row[COL_GRAIBAIT]),
                used_difron=_bool(row[COL_DIFRON]),
                used_fly_attractant=_bool(row[COL_FLY_ATTR]),
                # Keep generic fields empty for excel imports
                work_type='',
            )

            existing = FieldWorkOrder.objects.filter(order_number=order_num, source='excel').first()

            if existing:
                if update_existing:
                    for k, v in defaults.items():
                        setattr(existing, k, v)
                    existing.save()
                    updated += 1
                else:
                    skipped += 1
            else:
                FieldWorkOrder.objects.create(order_number=order_num, **defaults)
                created += 1

        self.stdout.write(
            f'  [{sheet_name}] created={created}  updated={updated}  skipped={skipped}'
        )
