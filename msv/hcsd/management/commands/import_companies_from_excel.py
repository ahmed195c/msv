from datetime import date, datetime
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError

from hcsd.models import Company, Enginer


DEFAULT_XLSX = "/home/a/Downloads/مزاولة-مهندس - SCM-AGE-ENV-F-85-01.xlsx"


class Command(BaseCommand):
    help = "Import companies and engineers from the provided Excel sheet."

    def add_arguments(self, parser):
        parser.add_argument("--file", default=DEFAULT_XLSX, help="Path to .xlsx file")
        parser.add_argument("--dry-run", action="store_true", help="Validate and print counters only")

    def handle(self, *args, **options):
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise CommandError("openpyxl is required. Install with: pip install openpyxl") from exc

        file_path = Path(options["file"]).expanduser()
        dry_run = bool(options["dry_run"])
        if not file_path.exists():
            raise CommandError(f"File not found: {file_path}")

        wb = load_workbook(file_path, data_only=True)
        ws = wb[wb.sheetnames[0]]

        # Arabic headers in row 1.
        header = [self._clean(v) for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        idx = {name: i for i, name in enumerate(header)}

        required = [
            "الاسم التجاري",
            "رقم الرخصة",
            "تاريخ انتهاء الرخصة",
            "الموقع",
        ]
        missing = [col for col in required if col not in idx]
        if missing:
            raise CommandError(f"Missing required columns in sheet: {', '.join(missing)}")

        counters = {
            "company_added": 0,
            "company_updated": 0,
            "company_skipped": 0,
            "engineer_added": 0,
            "engineer_updated": 0,
        }

        for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            company_name = self._clean(row[idx["الاسم التجاري"]])
            license_no = self._clean(row[idx["رقم الرخصة"]])
            license_exp = self._to_date(row[idx["تاريخ انتهاء الرخصة"]])
            address = self._clean(row[idx["الموقع"]])
            landline = self._clean(row[idx["رقم الهاتف الأرضي"]]) if "رقم الهاتف الأرضي" in idx else ""
            owner_phone = self._clean(row[idx["رقم هاتف المالك"]]) if "رقم هاتف المالك" in idx else ""
            company_email = self._clean_email(row[idx["البريد الالكتروني 1"]]) if "البريد الالكتروني 1" in idx else ""
            engineer_pairs = []
            id_col_1 = self._first_existing_header(
                idx,
                ["رقم الهوية / الرقم الموحد", "رقم الهوية", "الرقم الموحد", "الهوية الإماراتية"],
            )
            id_col_2 = self._first_existing_header(
                idx,
                ["رقم الهوية / الرقم الموحد (2)", "رقم الهوية (2)", "الرقم الموحد (2)", "الهوية الإماراتية (2)"],
            )
            email_col_1 = self._first_existing_header(
                idx,
                ["البريد الالكتروني للمهندس", "البريد الإلكتروني للمهندس", "ايميل المهندس"],
            )
            email_col_2 = self._first_existing_header(
                idx,
                ["البريد الالكتروني للمهندس (2)", "البريد الإلكتروني للمهندس (2)", "ايميل المهندس (2)"],
            )
            if "اسم المهندس" in idx:
                engineer_pairs.append(
                    (
                        self._clean(row[idx["اسم المهندس"]]),
                        self._clean(row[idx["رقم المتحرك"]]) if "رقم المتحرك" in idx else "",
                        self._clean(row[idx[id_col_1]]) if id_col_1 else "",
                        self._clean_email(row[idx[email_col_1]]) if email_col_1 else "",
                    )
                )
            if "اسم المهندس (2)" in idx:
                engineer_pairs.append(
                    (
                        self._clean(row[idx["اسم المهندس (2)"]]),
                        self._clean(row[idx["رقم المتحرك (2)"]]) if "رقم المتحرك (2)" in idx else "",
                        self._clean(row[idx[id_col_2]]) if id_col_2 else "",
                        self._clean_email(row[idx[email_col_2]]) if email_col_2 else "",
                    )
                )

            if not company_name or not license_no:
                counters["company_skipped"] += 1
                self.stdout.write(self.style.WARNING(f"Row {row_no}: skipped (missing company name/license)"))
                continue

            engineers = []
            for engineer_name, engineer_phone, engineer_id_number, engineer_email in engineer_pairs:
                engineer = self._upsert_engineer(
                    engineer_name=engineer_name,
                    engineer_phone=engineer_phone,
                    engineer_id_number=engineer_id_number,
                    engineer_email=engineer_email,
                    dry_run=dry_run,
                    counters=counters,
                )
                if engineer and engineer not in engineers:
                    engineers.append(engineer)
            primary_engineer = engineers[0] if engineers else None

            company = Company.objects.filter(number=license_no).first()
            if not company:
                counters["company_added"] += 1
                if dry_run:
                    continue
                company = Company.objects.create(
                    name=company_name,
                    number=license_no,
                    trade_license_exp=license_exp,
                    address=address or "-",
                    landline=landline or None,
                    owner_phone=owner_phone or None,
                    email=company_email or None,
                    business_activity="pest_control",
                    pest_control_type="public_health_pest_control",
                    enginer=primary_engineer,
                )
                if not dry_run and engineers:
                    company.engineers.set(engineers)
                continue

            changed = []
            if company.name != company_name:
                company.name = company_name
                changed.append("name")
            if license_exp and company.trade_license_exp != license_exp:
                company.trade_license_exp = license_exp
                changed.append("trade_license_exp")
            if address and company.address != address:
                company.address = address
                changed.append("address")
            if landline and company.landline != landline:
                company.landline = landline
                changed.append("landline")
            if owner_phone and company.owner_phone != owner_phone:
                company.owner_phone = owner_phone
                changed.append("owner_phone")
            if company_email and company.email != company_email:
                company.email = company_email
                changed.append("email")
            if company.business_activity != "pest_control":
                company.business_activity = "pest_control"
                changed.append("business_activity")
            if company.pest_control_type != "public_health_pest_control":
                company.pest_control_type = "public_health_pest_control"
                changed.append("pest_control_type")
            if primary_engineer and company.enginer_id != primary_engineer.id:
                company.enginer = primary_engineer
                changed.append("enginer")

            if changed:
                counters["company_updated"] += 1
                if not dry_run:
                    company.save(update_fields=changed)
                if not dry_run and engineers:
                    company.engineers.set(engineers)
            else:
                counters["company_skipped"] += 1
                if not dry_run and engineers:
                    company.engineers.set(engineers)

        self.stdout.write(self.style.SUCCESS("Import completed"))
        for key, value in counters.items():
            self.stdout.write(f"{key}: {value}")

    def _upsert_engineer(
        self,
        engineer_name,
        engineer_phone,
        engineer_id_number,
        engineer_email,
        dry_run,
        counters,
    ):
        if not engineer_name:
            return None
        normalized_email = engineer_email if engineer_email and self._is_valid_email(engineer_email) else ""
        engineer = None
        if engineer_id_number:
            engineer = Enginer.objects.filter(national_or_unified_number=engineer_id_number).first()
        if not engineer and engineer_phone:
            engineer = Enginer.objects.filter(phone=engineer_phone).first()
        if not engineer and engineer_email and self._is_valid_email(engineer_email):
            engineer = Enginer.objects.filter(email=engineer_email).first()
        if not engineer:
            engineer = Enginer.objects.filter(name=engineer_name).first()
        if not engineer:
            counters["engineer_added"] += 1
            if dry_run:
                return None
            return Enginer.objects.create(
                name=engineer_name,
                email=normalized_email,
                phone=engineer_phone or "0000000000",
                national_or_unified_number=engineer_id_number or None,
            )

        changed = []
        if engineer.name != engineer_name:
            engineer.name = engineer_name
            changed.append("name")
        if engineer_phone and engineer.phone != engineer_phone:
            engineer.phone = engineer_phone
            changed.append("phone")
        if engineer_id_number and engineer.national_or_unified_number != engineer_id_number:
            engineer.national_or_unified_number = engineer_id_number
            changed.append("national_or_unified_number")
        if normalized_email and engineer.email != normalized_email:
            engineer.email = normalized_email
            changed.append("email")
        if changed:
            counters["engineer_updated"] += 1
            if not dry_run:
                engineer.save(update_fields=changed)
        return engineer

    @staticmethod
    def _clean(value):
        if value is None:
            return ""
        return str(value).strip().strip("\t")

    @staticmethod
    def _clean_email(value):
        v = Command._clean(value).replace(" ", "")
        return v or ""

    @staticmethod
    def _to_date(value):
        if not value:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        return None

    @staticmethod
    def _is_valid_email(value):
        v = (value or "").strip()
        return bool(v and "@" in v and "." in v.split("@")[-1])

    @staticmethod
    def _first_existing_header(header_index, candidates):
        for col in candidates:
            if col in header_index:
                return col
        return None
