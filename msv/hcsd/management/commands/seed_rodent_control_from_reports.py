"""
One-time historical import: builds RodentControlBuilding + RodentControlVisit
records from the existing "Park RBS and Manholes Report.xlsx" spreadsheet
(the only one of the three reference reports with per-site, not per-area,
granularity). Safe to re-run — buildings are matched by name and visits are
upserted per (building, month), so re-running after fixing the source file
just updates existing records instead of duplicating them.

The "F-U" (Follow Up) sheet is a derived before/after comparison, not a
routine monthly visit log, and is intentionally skipped.
"""
import datetime

import openpyxl
from django.core.management.base import BaseCommand, CommandError

from hcsd.models import RodentControlBuilding, RodentControlVisit

DEFAULT_PATH = 'hcsd/static/hcsd/excl/buldings/Park RBS and Manholes Report.xlsx'

# sheet name -> (year, month). Determined from the actual "Date Of Services"
# values inside each sheet, not guessed from the sheet name alone.
MONTH_SHEETS = {
    'DEC': (2025, 12), 'Jan': (2026, 1), 'FEB': (2026, 2), 'MAR': (2026, 3),
    'APR': (2026, 4), 'MAY': (2026, 5), 'June': (2026, 6), 'July': (2026, 7),
    'AUG': (2026, 8), 'SEP': (2026, 9),
}

COUNT_FIELDS = {
    'Inspected RBS ( Park, Gov office )': 'inspected',
    'Infested RBS': 'infested',
    'Damages RBS': 'damaged',
    'Damage Lock': 'damaged',
    'New Installed RBS': 'newly_installed',
    'Replenished RBS': 'replenished',
}
NOTE_FIELDS = [
    'Inspected manhole', 'Infested manhole', 'Outside Burrows', 'Infested Burrows',
]


def _find_header_row(ws):
    for r in range(1, 4):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if 'Area Name' in vals:
            return r
    return None


def _rodenticide_name(header):
    # e.g. 'Rodenticide 1 {SUREFIRE ALL WEATHER.}  Qyt' -> 'SUREFIRE ALL WEATHER.'
    # or 'SUREFIRE ALL WEATHER WB (pcs)' -> 'SUREFIRE ALL WEATHER WB'
    if '{' in header and '}' in header:
        return header.split('{', 1)[1].split('}', 1)[0].strip()
    return header.replace('(pcs)', '').strip()


class Command(BaseCommand):
    help = 'One-time import of historical rodent-control data from Park RBS and Manholes Report.xlsx'

    def add_arguments(self, parser):
        parser.add_argument('--path', default=DEFAULT_PATH)
        parser.add_argument('--dry-run', action='store_true')

    def handle(self, *args, **options):
        path = options['path']
        dry_run = options['dry_run']

        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except FileNotFoundError:
            raise CommandError(f'File not found: {path}')

        buildings_created = 0
        visits_created = 0
        visits_updated = 0
        skipped_sheets = []

        for sheet_name, (year, month) in MONTH_SHEETS.items():
            if sheet_name not in wb.sheetnames:
                skipped_sheets.append(sheet_name)
                continue
            ws = wb[sheet_name]
            header_row = _find_header_row(ws)
            if header_row is None:
                skipped_sheets.append(sheet_name)
                continue

            headers = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]
            col_of = {h: i + 1 for i, h in enumerate(headers) if h}

            area_col = col_of.get('Area Name')
            date_col = col_of.get('Date Of Services')
            if not area_col:
                skipped_sheets.append(sheet_name)
                continue

            rodenticide_cols = [
                (c, _rodenticide_name(h)) for h, c in col_of.items()
                if h and ('Qyt' in h or 'Qty' in h or '(pcs)' in h)
            ]

            period_start = datetime.date(year, month, 1)

            for r in range(header_row + 1, ws.max_row + 1):
                name = ws.cell(row=r, column=area_col).value
                if not name or not str(name).strip():
                    continue
                name = str(name).strip()

                if dry_run:
                    building = RodentControlBuilding.objects.filter(name=name).first()
                else:
                    building, created = RodentControlBuilding.objects.get_or_create(name=name)
                    if created:
                        buildings_created += 1

                raw_date = ws.cell(row=r, column=date_col).value if date_col else None
                visit_date = raw_date.date() if isinstance(raw_date, datetime.datetime) else (
                    raw_date if isinstance(raw_date, datetime.date) else None
                )

                def _num(col):
                    v = ws.cell(row=r, column=col).value
                    try:
                        return float(v)
                    except (TypeError, ValueError):
                        return 0

                flags = {'inspected': False, 'infested': False, 'damaged': False,
                         'newly_installed': False, 'replenished': False}
                for header, field in COUNT_FIELDS.items():
                    c = col_of.get(header)
                    if c and _num(c) > 0:
                        flags[field] = True

                note_parts = []
                for header in NOTE_FIELDS:
                    c = col_of.get(header)
                    if c:
                        v = _num(c)
                        if v:
                            note_parts.append(f'{header}: {int(v) if v == int(v) else v}')

                rod_parts = []
                total_qty = 0.0
                for c, prod_name in rodenticide_cols:
                    v = _num(c)
                    if v:
                        rod_parts.append(f'{prod_name}: {v:g}')
                        total_qty += v

                notes = ' | '.join(note_parts)
                rodenticide_type = ', '.join(rod_parts)

                if dry_run:
                    continue

                if not building:
                    continue

                visit, created = RodentControlVisit.objects.update_or_create(
                    building=building, period_start=period_start,
                    defaults={
                        'visit_date': visit_date,
                        **flags,
                        'rodenticide_type': rodenticide_type,
                        'rodenticide_quantity': total_qty or None,
                        'notes': notes,
                    },
                )
                if created:
                    visits_created += 1
                else:
                    visits_updated += 1

        self.stdout.write(self.style.SUCCESS(
            f'Buildings created: {buildings_created}. '
            f'Visits created: {visits_created}, updated: {visits_updated}. '
            f'Skipped sheets: {skipped_sheets or "none"}.'
        ))
